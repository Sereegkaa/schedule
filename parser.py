# -*- coding: utf8 -*-
import os
import sqlite3
import time
from selenium import webdriver
from time import sleep
import urllib.request
import openpyxl

days = {'пн': 1, 'вт': 2, 'ср': 3, 'чт': 4, 'пт': 5, 'сб': 6}
keys = ['день недели', '№ пары', 'время', 'ауд', 'вид занятий', 'препод', 'пара', 'пара', 'препод', 'вид занятий',
        'ауд', '№ пары', 'время']
keys_bd = {'день недели': 'day', '№ пары': 'num_lesson', 'время': 'time', 'ауд': 'classroom', 'вид занятий': 'type',
           'препод': 'teacher', 'пара': 'lesson', 'группа': 'group', 'четная': 'chetn'}
con = sqlite3.connect("schedule.db")
cur = con.cursor()


def download():
    sp = ["курс", "живопись", "искусствоведение", "культурология", "лингвистика", " культура", "педагогическое",
          "филология", "хореография"]

    for filename in os.listdir("./files"):
        os.remove(os.path.join("./files", filename))

    driver = webdriver.Chrome(executable_path="chromedriver.exe")
    driver.get("https://kosygin-rgu.ru/1AppRGU/rguschedule/indexClassSchedule.aspx")
    sleep(2)

    for x in driver.find_elements_by_css_selector("div.panel-group"):
        if " очная " in x.text.lower():
            x.click()
            for y in x.find_elements_by_tag_name("table"):
                link = y.find_element_by_css_selector("td:last-child>a").get_attribute("href")
                print(link)
                urllib.request.urlretrieve(link, './files/' + link.split('/')[-1])


def parse_table():
    for sheetname in table.sheetnames:
        sheet = table[sheetname]
        group = ''
        day = ''
        if not (sheet['A1'].value or sheet['A2'].value or sheet['A3'].value):
            continue
        for row in range(2, sheet.max_row + 1):
            rasp_nech = {}
            rasp_ch = {}
            cell_1 = sheet.cell(row=row, column=1).value
            cell_2 = sheet.cell(row=row, column=2).value

            if not cell_1:
                if cell_2 in days.values():
                    cell_1 = day
                else:
                    continue
            if type(cell_1) == int:
                continue
            if group == '':
                if 'группа' in cell_1.lower():
                    group = cell_1.lower().replace('группа', '').split('(')[0].strip().replace(' ', '')
            elif cell_1.lower() in days.keys():
                day = cell_1.lower()
                rasp_nech['день недели'] = days[day]
                rasp_nech['четная'] = 0
                rasp_ch['день недели'] = days[day]
                rasp_ch['четная'] = 1
                rasp_ch['группа'] = group
                rasp_nech['группа'] = group
                for i in range(1, len(keys)):
                    if i <= 6:
                        rasp_nech[keys[i]] = (sheet.cell(row=row, column=i + 1).value or '-')
                    else:
                        rasp_ch[keys[i]] = (sheet.cell(row=row, column=i + 1).value or '-')
            if rasp_nech:
                req = f"""INSERT INTO schedule VALUES({rasp_nech['день недели']},'{rasp_nech['№ пары']}','{rasp_nech['время']}','{rasp_nech['ауд']}','{rasp_nech['вид занятий']}','{rasp_nech['препод']}',{rasp_nech['четная']},'{rasp_nech['пара']}','{rasp_nech['группа']}')"""
                # print(req)
                cur.execute(req)
                # time.sleep(2)

                req = f"""INSERT INTO schedule VALUES({rasp_nech['день недели']},'{rasp_ch['№ пары']}','{rasp_ch['время']}','{rasp_ch['ауд']}','{rasp_ch['вид занятий']}','{rasp_ch['препод']}',{rasp_ch['четная']},'{rasp_ch['пара']}','{rasp_ch['группа']}')"""
                # print(req)
                cur.execute(req)
                con.commit()
                # f.write(str(rasp_nech) + '\n')
                # f.write(str(rasp_ch) + '\n')
                # print(rasp_nech)
            # time.sleep(2)
        print(group)


start_time = time.process_time()
download()
tables = os.listdir('files')
for x in tables:
    table = openpyxl.load_workbook(f"files/{x}")
    print(f"------{x}------")
    parse_table()

print(time.process_time() - start_time, "seconds")
